#Python 3
#3/20/2020
#Rupak Kannan and Shyam Kannan
#Auto Valuation

import sys
import openpyxl
import csv
import pandas as pd
#Just a forewarning this module requires requests html
#yahoo_fin is currently broken for some unknown reason
from yahoo_fin import stock_info
import yfinance as yf
#from xlsxwriter.workbook import Workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from openpyxl.worksheet.datavalidation import DataValidation
import shutil
import os
import os.path
from openpyxl.utils import get_column_letter
from PyQt5 import QtWidgets, QtCore, QtGui
from PyQt5.QtWidgets import QMessageBox, QDialogButtonBox, QFileDialog



def write_to_target_page(target, directory, sheet_name):
    #writes to the target sheet page
    book = openpyxl.load_workbook(target)
    ws = book.get_sheet_by_name(sheet_name)
    f = open(directory)
    reader = csv.reader(f)
    #writes the values cell by cell to the selected page
    for row_index, row in enumerate(reader):
        for column_index, cell in enumerate(row):
            column_letter = get_column_letter((column_index + 1))
            s = cell
            try:
                try:
                    #remove the ',' in numbers
                    s=float(s.replace(',', ''))
                except:
                    #if number doesn't exist
                    s=float(s)
            except ValueError:
                pass

            ws.cell('%s%s'%(column_letter, (row_index + 1))).value = s

    book.save(filename = target)

def write_to_target_debt(target, directory, sheet_name):
    #writes the debt template to the debt template sheet (duh) if given
    wb1 = openpyxl.load_workbook(target)
    ws1 = wb1.get_sheet_by_name(sheet_name)

    wb2 = openpyxl.load_workbook(directory)
    ws2 = wb2.active
    mr = ws2.max_row
    mc = ws2.max_column

    for i in range(1, mr + 1):
        for j in range(1, mc + 1):
            c = ws2.cell(row = i, column = j)
            ws1.cell(row = i, column = j).value = c.value 
    wb1.save(filename = target)

#basic gui setup. Very messy but necessary for the gui to show up
class Ui_MainWindow(object):
    def setup_Ui(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(485, 660)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")

        self.title = QtWidgets.QLabel(self.centralwidget)
        self.title.setGeometry(QtCore.QRect(190, 40, 201, 51))
        font = QtGui.QFont()
        font.setFamily("Cambria")
        font.setPointSize(16)
        self.title.setFont(font)
        self.title.setObjectName("title")

        self.debt_group = QtWidgets.QButtonGroup()

        self.growth_rate = QtWidgets.QButtonGroup()

        self.logo = QtWidgets.QLabel(self.centralwidget)
        self.logo.setGeometry(QtCore.QRect(40, 10, 161, 111))
        self.logo.setText("")
        self.logo.setPixmap(QtGui.QPixmap("Capture.PNG"))
        self.logo.setObjectName("logo")

        #Main company info

        self.company_ticker_txt = QtWidgets.QLineEdit(self.centralwidget)
        self.company_ticker_txt.setGeometry(QtCore.QRect(190, 130, 61, 20))
        self.company_ticker_txt.setObjectName("company_ticker_txt")

        self.company_ticker_lbl = QtWidgets.QLabel(self.centralwidget)
        self.company_ticker_lbl.setGeometry(QtCore.QRect(40, 130, 131, 20))
        font = QtGui.QFont()
        font.setFamily("Cambria")
        font.setPointSize(10)
        self.company_ticker_lbl.setFont(font)
        self.company_ticker_lbl.setObjectName("company_ticket_lbl")

        self.company_ticker_btn = QtWidgets.QPushButton(self.centralwidget)
        self.company_ticker_btn.setGeometry(QtCore.QRect(245, 130, 125, 23))
        self.company_ticker_btn.setObjectName("company_ticker_btn")
        self.company_ticker_btn.clicked.connect(self.morningstar_download)

        self.balance_sheet_btn = QtWidgets.QPushButton(self.centralwidget)
        self.balance_sheet_btn.setGeometry(QtCore.QRect(390, 160, 75, 23))
        self.balance_sheet_btn.setObjectName("balance_sheet_btn")
        self.balance_sheet_btn.clicked.connect(self.get_xl_balance)

        self.balance_sheet_lbl = QtWidgets.QLabel(self.centralwidget)
        self.balance_sheet_lbl.setGeometry(QtCore.QRect(40, 160, 101, 20))
        font = QtGui.QFont()
        font.setFamily("Cambria")
        font.setPointSize(10)
        self.balance_sheet_lbl.setFont(font)
        self.balance_sheet_lbl.setObjectName("balance_sheet_lbl")

        self.balance_sheet_txt = QtWidgets.QLineEdit(self.centralwidget)
        self.balance_sheet_txt.setGeometry(QtCore.QRect(190, 160, 201, 21))
        self.balance_sheet_txt.setObjectName("balance_sheet_txt")
        self.balance_sheet_txt.setReadOnly(True)      


        self.cash_flow_txt = QtWidgets.QLineEdit(self.centralwidget)
        self.cash_flow_txt.setGeometry(QtCore.QRect(190, 190, 201, 21))
        self.cash_flow_txt.setObjectName("cash_flow_txt")
        self.cash_flow_txt.setReadOnly(True)        

        self.cash_flow_lbl = QtWidgets.QLabel(self.centralwidget)
        self.cash_flow_lbl.setGeometry(QtCore.QRect(40, 190, 101, 20))
        font = QtGui.QFont()
        font.setFamily("Cambria")
        font.setPointSize(10)
        self.cash_flow_lbl.setFont(font)
        self.cash_flow_lbl.setObjectName("cash_flow_lbl")

        self.cash_flow_btn = QtWidgets.QPushButton(self.centralwidget)
        self.cash_flow_btn.setGeometry(QtCore.QRect(390, 190, 75, 23))
        self.cash_flow_btn.setObjectName("cash_flow_btn")
        self.cash_flow_btn.clicked.connect(self.get_xl_cash)

        self.Income_statement_txt = QtWidgets.QLineEdit(self.centralwidget)
        self.Income_statement_txt.setGeometry(QtCore.QRect(190, 220, 201, 21))
        self.Income_statement_txt.setObjectName("Income_statement_txt")
        self.Income_statement_txt.setReadOnly(True)

        self.income_statement_btn = QtWidgets.QPushButton(self.centralwidget)
        self.income_statement_btn.setGeometry(QtCore.QRect(390, 220, 75, 23))
        self.income_statement_btn.setObjectName("income_statement_btn")
        self.income_statement_btn.clicked.connect(self.get_xl_income)

        self.income_statement_lbl = QtWidgets.QLabel(self.centralwidget)
        self.income_statement_lbl.setGeometry(QtCore.QRect(40, 220, 151, 20))
        font = QtGui.QFont()
        font.setFamily("Cambria")
        font.setPointSize(10)
        self.income_statement_lbl.setFont(font)
        self.income_statement_lbl.setObjectName("income_statement_lbl")

        self.debt_spreadsheet_lbl = QtWidgets.QLabel(self.centralwidget)
        self.debt_spreadsheet_lbl.setGeometry(QtCore.QRect(40, 280, 151, 20))
        font = QtGui.QFont()
        font.setFamily("Cambria")
        font.setPointSize(10)
        self.debt_spreadsheet_lbl.setFont(font)
        self.debt_spreadsheet_lbl.setObjectName("debt_spreadsheet_lbl")

        self.debt_spreadsheet_btn = QtWidgets.QPushButton(self.centralwidget)
        self.debt_spreadsheet_btn.setGeometry(QtCore.QRect(390, 340, 75, 23))
        self.debt_spreadsheet_btn.setObjectName("debt_spreadsheet_btn")
        self.debt_spreadsheet_btn.clicked.connect(self.get_xl_debt)

        self.debt_spreadsheet_txt = QtWidgets.QLineEdit(self.centralwidget)
        self.debt_spreadsheet_txt.setGeometry(QtCore.QRect(190, 340, 201, 21))
        self.debt_spreadsheet_txt.setObjectName("debt_spreadsheet_txt")
        self.debt_spreadsheet_txt.setReadOnly(True)

        self.Financial_statement_rd = QtWidgets.QRadioButton(self.centralwidget)
        self.Financial_statement_rd.setGeometry(QtCore.QRect(60, 310, 271, 17))
        font = QtGui.QFont()
        font.setFamily("Cambria")
        font.setPointSize(11)
        self.Financial_statement_rd.setFont(font)
        self.debt_group.addButton(self.Financial_statement_rd)
        self.Financial_statement_rd.setChecked(True)

        self.spreadsheet_rd = QtWidgets.QRadioButton(self.centralwidget)
        self.spreadsheet_rd.setGeometry(QtCore.QRect(60, 340, 231, 17))
        font = QtGui.QFont()
        font.setFamily("Cambria")
        font.setPointSize(11)
        self.spreadsheet_rd.setFont(font)
        self.debt_group.addButton(self.spreadsheet_rd)
        self.spreadsheet_rd.setObjectName("spreadsheet_rd")        

        self.key_ratios_lbl = QtWidgets.QLabel(self.centralwidget)
        self.key_ratios_lbl.setGeometry(QtCore.QRect(40, 250, 101, 20))
        font = QtGui.QFont()
        font.setFamily("Cambria")
        font.setPointSize(10)
        self.key_ratios_lbl.setFont(font)
        self.key_ratios_lbl.setObjectName("key_ratios_lbl")

        self.key_ratios_btn = QtWidgets.QPushButton(self.centralwidget)
        self.key_ratios_btn.setGeometry(QtCore.QRect(390, 250, 75, 23)) 
        self.key_ratios_btn.clicked.connect(self.get_xl_ratios)

        self.key_ratios_txt = QtWidgets.QLineEdit(self.centralwidget)
        self.key_ratios_txt.setGeometry(QtCore.QRect(190, 250, 201, 21))
        self.key_ratios_txt.setReadOnly(True)  

        #Additional financial info

        self.mrperp__txt = QtWidgets.QLineEdit(self.centralwidget)
        self.mrperp__txt.setGeometry(QtCore.QRect(190, 370, 61, 20))
        self.mrperp__txt.setText("")
        self.mrperp__txt.setObjectName("mrperp__txt")
        self.onlyNumbers = QtGui.QDoubleValidator()
        self.mrperp__txt.setValidator(self.onlyNumbers)     

        self.mrperp_lbl = QtWidgets.QLabel(self.centralwidget)
        self.mrperp_lbl.setGeometry(QtCore.QRect(40, 370, 101, 20))
        font = QtGui.QFont()
        font.setFamily("Cambria")
        font.setPointSize(10)
        self.mrperp_lbl.setFont(font)
        self.mrperp_lbl.setObjectName("mrperp_lbl")

        self.risk_free_rate_txt = QtWidgets.QLineEdit(self.centralwidget)
        self.risk_free_rate_txt.setGeometry(QtCore.QRect(190, 400, 61, 20))
        self.risk_free_rate_txt.setText("")
        self.risk_free_rate_txt.setObjectName("risk_free_rate_txt")
        self.onlyNumbers = QtGui.QDoubleValidator()
        self.risk_free_rate_txt.setValidator(self.onlyNumbers)        

        self.risk_free_rate_lbl = QtWidgets.QLabel(self.centralwidget)
        self.risk_free_rate_lbl.setGeometry(QtCore.QRect(40, 400, 141, 20))
        font = QtGui.QFont()
        font.setFamily("Cambria")
        font.setPointSize(10)
        self.risk_free_rate_lbl.setFont(font)
        self.risk_free_rate_lbl.setObjectName("risk_free_rate_lbl")

        self.terminal_txt = QtWidgets.QLineEdit(self.centralwidget)
        self.terminal_txt.setGeometry(QtCore.QRect(190, 430, 61, 20))
        self.terminal_txt.setText("")
        self.terminal_txt.setObjectName("terminal_txt")
        self.terminal_txt.setValidator(self.onlyNumbers)
        self.terminal_lbl = QtWidgets.QLabel(self.centralwidget)
        self.terminal_lbl.setGeometry(QtCore.QRect(10, 430, 170, 20))
        font = QtGui.QFont()
        font.setFamily("Cambria")
        font.setPointSize(10)
        self.terminal_lbl.setFont(font)
        self.terminal_lbl.setObjectName("terminal_lbl")

        self.year_growth_txt = QtWidgets.QSpinBox(self.centralwidget)
        self.year_growth_txt.setGeometry(QtCore.QRect(190, 460, 61, 20))
        self.year_growth_txt.setObjectName("risk_free_rate_txt")

        self.year_growth_lbl = QtWidgets.QLabel(self.centralwidget)
        self.year_growth_lbl.setGeometry(QtCore.QRect(10, 460, 170, 20))
        font = QtGui.QFont()
        font.setFamily("Cambria")
        font.setPointSize(10)
        self.year_growth_lbl.setFont(font)
        self.year_growth_lbl.setObjectName("risk_free_rate_lbl")
        self.year_growth_txt.setMinimum(0)
        self.year_growth_txt.setMaximum(10)

        self.year_growth_ddm_txt = QtWidgets.QSpinBox(self.centralwidget)
        self.year_growth_ddm_txt.setGeometry(QtCore.QRect(190, 490, 61, 20))
        self.year_growth_ddm_txt.setObjectName("risk_free_rate_txt")        

        self.year_growth_ddm_lbl = QtWidgets.QLabel(self.centralwidget)
        self.year_growth_ddm_lbl.setGeometry(QtCore.QRect(10, 490, 170, 20))
        font = QtGui.QFont()
        font.setFamily("Cambria")
        font.setPointSize(8)
        self.year_growth_ddm_lbl.setFont(font)
        self.year_growth_ddm_lbl.setObjectName("risk_free_rate_lbl")
        self.year_growth_ddm_txt.setMinimum(0)
        self.year_growth_ddm_txt.setMaximum(10)        

        self.growth_rate_lbl = QtWidgets.QLabel(self.centralwidget)
        self.growth_rate_lbl.setGeometry(QtCore.QRect(40, 520, 130, 20))
        font = QtGui.QFont()
        font.setFamily("Cambria")
        font.setPointSize(10)
        self.growth_rate_lbl.setFont(font)
        self.growth_rate_lbl.setObjectName("growth_rate_lbl")

        self.smallest_of_etc_rd = QtWidgets.QRadioButton(self.centralwidget)
        self.smallest_of_etc_rd.setGeometry(QtCore.QRect(60, 550, 271, 17))
        font = QtGui.QFont()
        font.setFamily("Cambria")
        font.setPointSize(11)
        self.smallest_of_etc_rd.setFont(font)
        self.growth_rate.addButton(self.smallest_of_etc_rd)
        self.smallest_of_etc_rd.setChecked(True)

        self.custom_rd = QtWidgets.QRadioButton(self.centralwidget)
        self.custom_rd.setGeometry(QtCore.QRect(60, 580, 231, 17))
        font = QtGui.QFont()
        font.setFamily("Cambria")
        font.setPointSize(11)
        self.custom_rd.setFont(font)
        self.growth_rate.addButton(self.custom_rd)
        self.custom_rd.setObjectName("custom_rd")

        self.custom_txt = QtWidgets.QLineEdit(self.centralwidget)
        self.custom_txt.setGeometry(QtCore.QRect(160, 580, 51, 20))
        self.custom_txt.setObjectName("custom_txt")
        self.onlyNumbers = QtGui.QDoubleValidator()
        self.custom_txt.setValidator(self.onlyNumbers)         

        self.run_btn = QtWidgets.QPushButton(self.centralwidget)
        self.run_btn.setGeometry(QtCore.QRect(50, 610, 75, 23))
        self.run_btn.setObjectName("run_btn")
        self.run_btn.clicked.connect(self.run)

        self.reset_btn = QtWidgets.QPushButton(self.centralwidget)
        self.reset_btn.setGeometry(QtCore.QRect(140, 610, 75, 23))
        self.reset_btn.setObjectName("reset_btn")
        self.reset_btn.clicked.connect(self.reset)

        self.close_btn = QtWidgets.QPushButton(self.centralwidget)
        self.close_btn.setGeometry(QtCore.QRect(230, 610, 75, 23))
        self.close_btn.setObjectName("close_btn")
        self.close_btn.clicked.connect(self.close)

        self.get_bond_data = QtWidgets.QPushButton(self.centralwidget)
        self.get_bond_data.setGeometry(QtCore.QRect(320, 610, 135, 23))
        self.get_bond_data.clicked.connect(self.open_bond)


        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 580, 21))
        self.menubar.setObjectName("menubar")

        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def morningstar_download(self):

        ticker = self.company_ticker_txt.text()
        location = os.getcwd()
        path = os.path.join(location, ticker)

        try:
            shutil.rmtree(path)
        except:
            pass

        try:
            options = webdriver.ChromeOptions()
            current_dir = os.getcwd()
            prefs = {
                "download.default_directory": current_dir + '\\' + ticker,
            "download.prompt_for_download": False,
            "download.directory_upgrade": True
            }            
            options.add_experimental_option('prefs', prefs)
            options.add_argument("headless")
            ticker = self.company_ticker_txt.text()   
            driver = webdriver.Chrome(ChromeDriverManager().install(), chrome_options = options)
            driver.set_window_size(777, 777)
            driver.get("http://financials.morningstar.com/balance-sheet/bs.html?t="+ticker+"&region=usa&culture=en-US")
            driver.execute_script("javascript:SRT_stocFund.orderControl('desc','Descending')")
            driver.execute_script("javascript:SRT_stocFund.Export();")
            driver.get("http://financials.morningstar.com/income-statement/is.html?t="+ticker+"&region=usa&culture=en-US")
            driver.execute_script("javascript:SRT_stocFund.orderControl('desc','Descending')")
            driver.execute_script("javascript:SRT_stocFund.Export();")
            driver.get("http://financials.morningstar.com/cash-flow/cf.html?t="+ticker+"&region=usa&culture=en-US")
            driver.execute_script("javascript:SRT_stocFund.orderControl('desc','Descending')")
            driver.execute_script("javascript:SRT_stocFund.Export();")
            driver.get("http://financials.morningstar.com/ratios/r.html?t="+ticker+"&region=usa&culture=en-US")
            driver.execute_script("javascript:orderChange('desc','Descending');")
            driver.execute_script("javascript:exportKeyStat2CSV();")
            driver.get("http://www.google.com")
            driver.quit()

            #Adds files to the txt statement

            self.Income_statement_txt.setText(os.path.realpath(ticker + "\\{} Income Statement.csv").format(ticker) )            
            self.balance_sheet_txt.setText(os.path.realpath(ticker + "\\{} Balance Sheet.csv").format(ticker))           
            self.cash_flow_txt.setText(os.path.realpath(ticker + "\\{} Cash Flow.csv").format(ticker))            
            self.key_ratios_txt.setText(os.path.realpath(ticker + "\\{} Key Ratios.csv").format(ticker))     


        except:
            try:
                driver.quit()
            except:
                pass

            msg = QMessageBox()
            msg.setWindowTitle("Notice")
            msg.setIcon(QMessageBox.Information)
            msg.setText("An error has occured, there is either: \n* An nonexistent company ticker \n* Error on Morningstar's website  \n* Interruption with the download process \nPlease try again.")
            notice = msg.exec()

    def get_xl_income(self):
        income_filename, filter = QtWidgets.QFileDialog.getOpenFileName(caption='Open file',  filter='CSV (*.CSV)')

        if income_filename:
            self.Income_statement_txt.setText(income_filename)

    def get_xl_balance(self):
        balance_filename, filter = QtWidgets.QFileDialog.getOpenFileName(caption='Open file',  filter='CSV (*.CSV)')

        if balance_filename:
            self.balance_sheet_txt.setText(balance_filename)         

    def get_xl_cash(self):
        cash_filename, filter = QtWidgets.QFileDialog.getOpenFileName(caption='Open file',  filter='CSV (*.CSV)')

        if cash_filename:
            self.cash_flow_txt.setText(cash_filename)

    def get_xl_debt(self):
        debt_filename, filter = QtWidgets.QFileDialog.getOpenFileName(caption='Open file',  filter='xlsx (*.xlsx)')

        if debt_filename:
            self.debt_spreadsheet_txt.setText(debt_filename)             


    def get_xl_ratios(self):
        ratios_filename, filter = QtWidgets.QFileDialog.getOpenFileName(caption='Open file',  filter='CSV (*.CSV)')

        if ratios_filename:
            self.key_ratios_txt.setText(ratios_filename)  

    def run(self):
        if self.Income_statement_txt.text() == "" or  self.balance_sheet_txt.text() == "" or self.cash_flow_txt.text() == "" or self.key_ratios_txt.text() == "" or self.company_ticker_txt.text() == "" or self.mrperp__txt.text() == "" or self.risk_free_rate_txt.text() == "":
            msg = QMessageBox()
            msg.setWindowTitle("Notice")
            msg.setIcon(QMessageBox.Information)
            msg.setText("All information must be filled")
            notice = msg.exec()   

        if self.spreadsheet_rd.isChecked() and self.debt_spreadsheet_txt.text() == "":
            msg = QMessageBox()
            msg.setWindowTitle("Notice")
            msg.setIcon(QMessageBox.Information)
            msg.setText("All information must be filled")
            notice = msg.exec()               

        if self.custom_rd.isChecked() and self.custom_txt.text() == "":
            msg = QMessageBox()
            msg.setWindowTitle("Notice")
            msg.setIcon(QMessageBox.Information)
            msg.setText("All information must be filled")
            notice = msg.exec()            

        else:
            try:
                #once the user confirms the info in the text is what he wants, start writing to excel (important)
                ticker = self.company_ticker_txt.text()
                original = 'Template.xlsx'
                target = ticker + '_Valuation.xlsx'
                shutil.copyfile(original, target)

                book = openpyxl.load_workbook(target, data_only=False)

                ws = book.get_sheet_by_name('DCF')

                ws['B7'].value = int(self.year_growth_txt.value())

                terminal_decimal = float(self.terminal_txt.text()) / 100
                ws['B8'].value = float(terminal_decimal)
                ws['B8'].number_format = '0.00%'

                risk_free_decimal = float(self.risk_free_rate_txt.text()) / 100
                ws['P17'].value = float(risk_free_decimal)
                ws['P17'].number_format = '0.00%'  

                MRP_decimal = float(self.mrperp__txt.text()) / 100
                ws['P9'].value = float(MRP_decimal)
                ws['P9'].number_format = '0.00%'

                ws['P12'].value = stock_info.get_live_price(ticker)

                beta = stock_info.get_quote_table(ticker, dict_result=True)
                ws['P2'].value = beta['Beta (5Y Monthly)']

                ws = book.get_sheet_by_name('Growth Rates')

                if self.custom_rd.isChecked():
                    custom_decimal = float(self.custom_txt.text()) / 100
                    ws['B18'].value = float(custom_decimal)


                if self.Financial_statement_rd.isChecked():
                    ws = book.get_sheet_by_name('DCF')
                    ws['Y1'].value = 'Financial Statements'

                ws = book.get_sheet_by_name('DDM')

                ws['B5'].value = int(self.year_growth_ddm_txt.value())
                
                ws = book.get_sheet_by_name('Multiples')
                #valuation_ratios = stock_info.get_stats_valuation(ticker)
                valuation_ratios = yf.download(ticker)
                valuation_ratios.to_dict()
                ws['H12'].value = valuation_ratios.iat[3,1]
                ws['H13'].value = valuation_ratios.iat[6,1]
                ws['H14'].value = valuation_ratios.iat[5,1]
                ws['H15'].value = valuation_ratios.iat[8,1]
                
                ws = book.get_sheet_by_name('DCF')

                book.save(filename = target)

                write_to_target_page(target, self.Income_statement_txt.text(), 'Income Statement')
                write_to_target_page(target, self.balance_sheet_txt.text(), 'Balance Sheet (Annual)')
                write_to_target_page(target, self.cash_flow_txt.text(), 'Cash Flow Statement')
                write_to_target_page(target, self.key_ratios_txt.text(), 'Key Ratios')

                if self.spreadsheet_rd.isChecked():
                    write_to_target_debt(target, self.debt_spreadsheet_txt.text(), 'Debt Template')                        

                msg = QMessageBox()
                msg.setWindowTitle("Notice")
                msg.setIcon(QMessageBox.Information)
                msg.setText("Workbook Created!")
                notice = msg.exec()
                os.system("start EXCEL.EXE " + target)

            except Exception as e:
                msg = QMessageBox()
                msg.setWindowTitle("Notice")
                msg.setIcon(QMessageBox.Information)
                msg.setText("Something went wrong. Try redownloading the sheets form MorningStar or check the values that have been entered or check if an valuation excel is currently open and close it.")
                notice = msg.exec()
                print(e)
                

    def reset(self):
        self.Income_statement_txt.setText("")
        self.balance_sheet_txt.setText("")
        self.cash_flow_txt.setText("")
        self.debt_spreadsheet_txt.setText("")
        self.key_ratios_txt.setText("")
        self.company_ticker_txt.setText("")
        self.mrperp__txt.setText("")
        self.risk_free_rate_txt.setText("")
        self.terminal_txt.setText("")
        self.year_growth_txt.setValue(0)
        self.custom_txt.setText("")

    def close(self):
        app.quit()

    def open_bond(self):
        ticker = self.company_ticker_txt.text()
        original = 'Debt_Template.xlsx'
        target = ticker + '_Debt_Template.xlsx'
        shutil.copyfile(original, target)
        os.system("start EXCEL.EXE " + target)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "Valuation Builder ver.1.0"))
        self.title.setText(_translate("MainWindow", "Valuation Builder"))
        self.income_statement_btn.setText(_translate("MainWindow", "Choose File"))
        self.income_statement_lbl.setText(_translate("MainWindow", "Income Statment"))
        self.balance_sheet_btn.setText(_translate("MainWindow", "Choose File"))
        self.balance_sheet_lbl.setText(_translate("MainWindow", "Balance Sheet"))
        self.cash_flow_btn.setText(_translate("MainWindow", "Choose File"))
        self.cash_flow_lbl.setText(_translate("MainWindow", "Cash Flow"))
        self.company_ticker_lbl.setText(_translate("MainWindow", "Company Ticker"))
        self.company_ticker_btn.setText(_translate("MainWindow", "Get Financial Data"))
        self.debt_spreadsheet_lbl.setText(_translate("MainWindow", "Debt Source:"))
        self.debt_spreadsheet_btn.setText(_translate("MainWindow", "Choose File"))
        self.spreadsheet_rd.setText(_translate("MainWindow", "Spreadsheet:"))
        self.Financial_statement_rd.setText(_translate("MainWindow", "Financial Statement"))
        self.key_ratios_lbl.setText(_translate("MainWindow", "Key Ratios"))
        self.key_ratios_btn.setText(_translate("MainWindow", "Choose File"))
        self.mrperp_lbl.setText(_translate("MainWindow", "MRP/ERP"))
        self.risk_free_rate_lbl.setText(_translate("MainWindow", "Risk Free Rate"))
        self.terminal_lbl.setText(_translate("MainWindow", "Terminal Growth Rate"))
        self.year_growth_lbl.setText(_translate("MainWindow", "Years of High Growth"))
        self.year_growth_ddm_lbl.setText(_translate("MainWindow", "Years of High Growth (DDM)"))
        self.growth_rate_lbl.setText(_translate("MainWindow", "Growth Rate:"))
        self.smallest_of_etc_rd.setText(_translate("MainWindow", "Use smallest of IGR, SGR, or ROI"))
        self.custom_rd.setText(_translate("MainWindow", "Custom"))
        self.run_btn.setText(_translate("MainWindow", "Run"))
        self.reset_btn.setText(_translate("MainWindow", "Reset"))
        self.close_btn.setText(_translate("MainWindow", "Close"))
        self.get_bond_data.setText(_translate("MainWindow", "Create Bond Template"))

if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    app.setStyle('Windows')
    Gui = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setup_Ui(Gui)
    Gui.show()
    msg = QMessageBox()
    msg.setWindowTitle("Disclaimer")
    msg.setIcon(QMessageBox.Information)
    msg.setText("The morningstar website is a little unstable. Downloads may not work sometimes. Retry the download if you get an Error.")
    notice = msg.exec()    
    sys.exit(app.exec_())
