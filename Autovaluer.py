#Python 3
#3/20/2020
#Rupak Kannan and Shyam Kannan
#Auto Valuation

import sys
import openpyxl
from openpyxl import *
import shutil
import os
from PyQt5 import QtWidgets, QtCore, QtGui
from PyQt5.QtWidgets import QMessageBox, QDialogButtonBox, QFileDialog

book = openpyxl.load_workbook('Template.xlsx')
sheet = book.active

#openpyxl books
book_income = ''
book_balance = ''
book_cash = ''
book_debt = '' 
book_ratios = ''

class Ui_MainWindow(object):
    def setup_Ui(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(485, 500)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        
        self.title = QtWidgets.QLabel(self.centralwidget)
        self.title.setGeometry(QtCore.QRect(190, 40, 201, 51))
        font = QtGui.QFont()
        font.setFamily("Cambria")
        font.setPointSize(16)
        self.title.setFont(font)
        self.title.setObjectName("title")
        
        self.logo = QtWidgets.QLabel(self.centralwidget)
        self.logo.setGeometry(QtCore.QRect(40, 10, 161, 111))
        self.logo.setText("")
        self.logo.setPixmap(QtGui.QPixmap("Capture.PNG"))
        self.logo.setObjectName("logo")
        
        self.company_ticker_txt = QtWidgets.QLineEdit(self.centralwidget)
        self.company_ticker_txt.setGeometry(QtCore.QRect(180, 140, 61, 20))
        self.company_ticker_txt.setObjectName("company_ticker_txt")
        
        self.company_ticket_lbl = QtWidgets.QLabel(self.centralwidget)
        self.company_ticket_lbl.setGeometry(QtCore.QRect(60, 140, 101, 20))
        font = QtGui.QFont()
        font.setFamily("Cambria")
        font.setPointSize(10)
        self.company_ticket_lbl.setFont(font)
        self.company_ticket_lbl.setObjectName("company_ticket_lbl")
        
        self.debt_spreadsheet_lbl = QtWidgets.QLabel(self.centralwidget)
        self.debt_spreadsheet_lbl.setGeometry(QtCore.QRect(60, 170, 101, 20))
        font = QtGui.QFont()
        font.setFamily("Cambria")
        font.setPointSize(10)
        self.debt_spreadsheet_lbl.setFont(font)
        self.debt_spreadsheet_lbl.setObjectName("debt_spreadsheet_lbl")
        
        self.debt_spreadsheet_btn = QtWidgets.QPushButton(self.centralwidget)
        self.debt_spreadsheet_btn.setGeometry(QtCore.QRect(380, 170, 75, 23))
        self.debt_spreadsheet_btn.setObjectName("debt_spreadsheet_btn")
        self.debt_spreadsheet_btn.clicked.connect(self.get_xl_debt)
        
        self.debt_spreadsheet_txt = QtWidgets.QLineEdit(self.centralwidget)
        self.debt_spreadsheet_txt.setGeometry(QtCore.QRect(180, 170, 201, 21))
        self.debt_spreadsheet_txt.setObjectName("debt_spreadsheet_txt")
        self.debt_spreadsheet_txt.setReadOnly(True)
        
        self.mrperp__txt = QtWidgets.QLineEdit(self.centralwidget)
        self.mrperp__txt.setGeometry(QtCore.QRect(180, 200, 61, 20))
        self.mrperp__txt.setText("")
        self.mrperp__txt.setObjectName("mrperp__txt")
        
        self.mrperp_lbl = QtWidgets.QLabel(self.centralwidget)
        self.mrperp_lbl.setGeometry(QtCore.QRect(100, 200, 61, 20))
        font = QtGui.QFont()
        font.setFamily("Cambria")
        font.setPointSize(10)
        self.mrperp_lbl.setFont(font)
        self.mrperp_lbl.setObjectName("mrperp_lbl")
        
        self.risk_free_rate_txt = QtWidgets.QLineEdit(self.centralwidget)
        self.risk_free_rate_txt.setGeometry(QtCore.QRect(180, 230, 61, 20))
        self.risk_free_rate_txt.setText("")
        self.risk_free_rate_txt.setObjectName("risk_free_rate_txt")
        self.onlyNumbers = QtGui.QDoubleValidator()
        self.risk_free_rate_txt.setValidator(self.onlyNumbers)        
        
        self.risk_free_rate_lbl = QtWidgets.QLabel(self.centralwidget)
        self.risk_free_rate_lbl.setGeometry(QtCore.QRect(80, 230, 81, 20))
        font = QtGui.QFont()
        font.setFamily("Cambria")
        font.setPointSize(10)
        self.risk_free_rate_lbl.setFont(font)
        self.risk_free_rate_lbl.setObjectName("risk_free_rate_lbl")
        
        self.terminal_txt = QtWidgets.QLineEdit(self.centralwidget)
        self.terminal_txt.setGeometry(QtCore.QRect(180, 260, 61, 20))
        self.terminal_txt.setText("")
        self.terminal_txt.setObjectName("terminal_txt")
        self.terminal_txt.setValidator(self.onlyNumbers)         
        
        self.terminal_lbl = QtWidgets.QLabel(self.centralwidget)
        self.terminal_lbl.setGeometry(QtCore.QRect(35, 260, 150, 20))
        font = QtGui.QFont()
        font.setFamily("Cambria")
        font.setPointSize(10)
        self.terminal_lbl.setFont(font)
        self.terminal_lbl.setObjectName("terminal_lbl")
        
        self.year_growth_txt = QtWidgets.QLineEdit(self.centralwidget)
        self.year_growth_txt.setGeometry(QtCore.QRect(180, 290, 61, 20))
        self.year_growth_txt.setText("")
        self.year_growth_txt.setObjectName("risk_free_rate_txt")
        
        self.year_growth_lbl = QtWidgets.QLabel(self.centralwidget)
        self.year_growth_lbl.setGeometry(QtCore.QRect(40, 290, 150, 20))
        font = QtGui.QFont()
        font.setFamily("Cambria")
        font.setPointSize(10)
        self.year_growth_lbl.setFont(font)
        self.year_growth_lbl.setObjectName("risk_free_rate_lbl")
               
        self.growth_rate_lbl = QtWidgets.QLabel(self.centralwidget)
        self.growth_rate_lbl.setGeometry(QtCore.QRect(30, 330, 81, 20))
        font = QtGui.QFont()
        font.setFamily("Cambria")
        font.setPointSize(10)
        self.growth_rate_lbl.setFont(font)
        self.growth_rate_lbl.setObjectName("growth_rate_lbl")
        
        self.smallest_of_etc_rd = QtWidgets.QRadioButton(self.centralwidget)
        self.smallest_of_etc_rd.setGeometry(QtCore.QRect(60, 360, 231, 17))
        font = QtGui.QFont()
        font.setFamily("Cambria")
        font.setPointSize(11)
        self.smallest_of_etc_rd.setFont(font)
        self.smallest_of_etc_rd.setObjectName("smallest_of_etc_rd")
        
        self.custom_rd = QtWidgets.QRadioButton(self.centralwidget)
        self.custom_rd.setGeometry(QtCore.QRect(60, 390, 231, 17))
        font = QtGui.QFont()
        font.setFamily("Cambria")
        font.setPointSize(11)
        self.custom_rd.setFont(font)
        self.custom_rd.setObjectName("custom_rd")
        
        self.custom_txt = QtWidgets.QLineEdit(self.centralwidget)
        self.custom_txt.setGeometry(QtCore.QRect(140, 390, 51, 20))
        self.custom_txt.setObjectName("custom_txt")
        
        self.run_btn = QtWidgets.QPushButton(self.centralwidget)
        self.run_btn.setGeometry(QtCore.QRect(50, 420, 75, 23))
        self.run_btn.setObjectName("run_btn")
        self.run_btn.clicked.connect(self.run)
        
        self.reset_btn = QtWidgets.QPushButton(self.centralwidget)
        self.reset_btn.setGeometry(QtCore.QRect(140, 420, 75, 23))
        self.reset_btn.setObjectName("reset_btn")
        self.reset_btn.clicked.connect(self.reset)
        
        self.close_btn = QtWidgets.QPushButton(self.centralwidget)
        self.close_btn.setGeometry(QtCore.QRect(320, 420, 75, 23))
        self.close_btn.setObjectName("close_btn")
        self.close_btn.clicked.connect(self.close)
        
        self.get_bond_data = QtWidgets.QPushButton(self.centralwidget)
        self.get_bond_data.setGeometry(QtCore.QRect(230, 420, 75, 23))
        #self.get_bond_data.clicked.connect(self.close)
        
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 580, 21))
        self.menubar.setObjectName("menubar")
        
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)
    
    def get_xl_income(self):
        global book_income
        income_filename, filter = QtWidgets.QFileDialog.getOpenFileName(caption='Open file',  filter='CSV (*.CSV);;xlsx (*.xlsx)')

        if income_filename:
            self.Income_statement_txt.setText(income_filename)
            book_income = openpyxl.load_workbook(income_filename)
            sheet_income = book_income.active
        
    def get_xl_balance(self):
        global book_balance
        balance_filename, filter = QtWidgets.QFileDialog.getOpenFileName(caption='Open file',  filter='CSV (*.CSV);;xlsx (*.xlsx)')

        if balance_filename:
            self.balance_sheet_txt.setText(balance_filename)
            book_balance = openpyxl.load_workbook(balance_filename)
            sheet_balance = book_balance.active            
        
    def get_xl_cash(self):
        global book_cash
        cash_filename, filter = QtWidgets.QFileDialog.getOpenFileName(caption='Open file',  filter='CSV (*.CSV);;xlsx (*.xlsx)')

        if cash_filename:
            self.cash_file_txt.setText(cash_filename)
            book_cash = openpyxl.load_workbook(cash_filename)
            sheet_cash = book_cash.active
        
    def get_xl_debt(self):
        global book_debt
        debt_filename, filter = QtWidgets.QFileDialog.getOpenFileName(caption='Open file',  filter='xlsx (*.xlsx);;CSV (*.CSV)')

        if debt_filename:
            self.debt_spreadsheet_txt.setText(debt_filename) 
            book_debt = openpyxl.load_workbook(debt_filename)
            sheet_debt = book_debt.active            
            
    
    def get_xl_ratios(self):
        global book_ratios
        ratios_filename, filter = QtWidgets.QFileDialog.getOpenFileName(caption='Open file',  filter='CSV (*.CSV);;xlsx (*.xlsx)')

        if ratios_filename:
            self.key_ratios_txt.setText(ratios_filename)  
            book_ratios = openpyxl.load_workbook(ratios_filename)
            sheet_ratios = book_ratios.active            
    
    def run(self):
        if self.Income_statement_txt.text() == "" or  self.balance_sheet_txt.text() == "" or self.cash_file_txt.text() == "" or self.debt_spreadsheet_txt.text() == "" or self.key_ratios_txt.text() == "" or self.company_ticker_txt.text() == "" or self.mrperp__txt.text() == "" or self.risk_free_rate_txt.text() == "":
            msg = QMessageBox()
            msg.setWindowTitle("Notice")
            msg.setIcon(QMessageBox.Information)
            msg.setText("All information must be filled")
            notice = msg.exec()        
        
        elif self.custom_rd.isChecked() and self.custom_txt.text() == "":
            msg = QMessageBox()
            msg.setWindowTitle("Notice")
            msg.setIcon(QMessageBox.Information)
            msg.setText("All information must be filled")
            notice = msg.exec()            

        else:
            ticker = self.company_ticker_txt.text()
            original = 'Template.xlsx'
            target = ticker + '_Valuation.xlsx'
            shutil.copyfile(original, target)
            msg = QMessageBox()
            msg.setWindowTitle("Notice")
            msg.setIcon(QMessageBox.Information)
            msg.setText("Workbook Created!")
            notice = msg.exec()
            os.system("start EXCEL.EXE " + target)
            book = openpyxl.load_workbook(target)
            book_income = openpyxl.load_workbook(target)
            
            
            
    def reset(self):
        self.Income_statement_txt.setText("")
        self.balance_sheet_txt.setText("")
        self.cash_file_txt.setText("")
        self.debt_spreadsheet_txt.setText("")
        self.key_ratios_txt.setText("")
        self.company_ticker_txt.setText("")
        self.mrperp__txt.setText("")
        self.risk_free_rate_txt.setText("")
        self.custom_txt.setText("")
        
    def close(self):
        app.quit()

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "Autovaluer"))
        self.title.setText(_translate("MainWindow", "Autovaluer"))
        self.company_ticket_lbl.setText(_translate("MainWindow", "Company Ticker"))
        self.debt_spreadsheet_lbl.setText(_translate("MainWindow", "Debt Spreadsheet"))
        self.debt_spreadsheet_btn.setText(_translate("MainWindow", "Choose File"))
        self.mrperp_lbl.setText(_translate("MainWindow", "MRP/ERP"))
        self.risk_free_rate_lbl.setText(_translate("MainWindow", "Risk Free Rate"))
        self.terminal_lbl.setText(_translate("MainWindow", "Terminal Growth Rate"))
        self.year_growth_lbl.setText(_translate("MainWindow", "Years of High Growth"))
        self.growth_rate_lbl.setText(_translate("MainWindow", "Growth Rate:"))
        self.smallest_of_etc_rd.setText(_translate("MainWindow", "Use smallest of IAR, SAR, or ROLC"))
        self.custom_rd.setText(_translate("MainWindow", "Custom"))
        self.run_btn.setText(_translate("MainWindow", "Run"))
        self.reset_btn.setText(_translate("MainWindow", "Reset"))
        self.close_btn.setText(_translate("MainWindow", "Close"))
        self.get_bond_data.setText(_translate("MainWindow", "Get Bond Data"))
        
if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    app.setStyle('Windows')
    Gui = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setup_Ui(Gui)
    Gui.show()  
    sys.exit(app.exec_())
